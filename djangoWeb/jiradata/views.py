from django.shortcuts import render, HttpResponse
from django.http import StreamingHttpResponse
from jira import JIRA, JIRAError
import json, time
from openpyxl import Workbook, load_workbook
from openpyxl.compat import range
from openpyxl.styles import Font, Alignment
from django.utils.http import urlquote
from django.conf import settings


# Create your views here.
def search_data(request):
    if request.method == 'POST':
        username = request.POST.get('JIRAId')
        code = request.POST.get('JIRACode')
        projectId = request.POST.get('ProjectId')
        model = request.POST.get('Model')
        Model = ''
        for item in model.split(';'):
            Model = Model + '产品机型~ "' + item + '" OR '
        result = {
            'success': False,
            'message': '',
            'data': [],
            'downloadId': ''
        }
        options = {
            'server': 'http://tvjira.hisense.com/'}
        try:
            jira_proj = JIRA(options, basic_auth=(username, code))
            if model == "":
                issues_in_proj = jira_proj.search_issues('project = ' + projectId + ' AND '
                                                                           '(issuetype in ("CTA Bug","OA Bug", "SW Bug", "SWD Bug ", "UE Bug") OR '
                                                                           '(issuetype = "TE Bug " AND ("缺陷来源(TE)" in (软件)))OR (issuetype = "NC Bug" AND NC类别 = 软件NC)) AND '
                                                                           'priority in (Safety, Critical, Major, Normal, Minor) AND status in (Reopened, Pending, Validating, Resolved, Open) '
                                                                           'ORDER BY priority DESC, created DESC', maxResults=100000)
            else:
                issues_in_proj = jira_proj.search_issues('project = ' + projectId + ' AND '
                                                         '(' + Model[:-4] +') AND '
                                                         '(issuetype in ("CTA Bug","OA Bug", "SW Bug", "SWD Bug ", "UE Bug") OR '
                                                         '(issuetype = "TE Bug " AND ("缺陷来源(TE)" in (软件)))OR (issuetype = "NC Bug" AND NC类别 = 软件NC)) AND '
                                                         'priority in (Safety, Critical, Major, Normal, Minor) AND status in (Reopened, Pending, Validating, Resolved, Open) '
                                                         'ORDER BY priority DESC, created DESC', maxResults=100000)
            store_list = []     # 运营问题列表
            ue_list = []        # UE问题列表
            for issue in issues_in_proj:
                # 遗留问题列表
                result_data = {}
                result_data['项目'] = projectId
                result_data['问题类型'] = str(issue.fields.issuetype)
                result_data['标识'] = issue.key
                result_data['主题'] = issue.fields.summary
                result_data['经办人'] = issue.fields.assignee.displayName
                result_data['报告人'] = issue.fields.reporter.displayName
                result_data['优先级'] = issue.fields.priority.name
                result_data['状态'] = issue.fields.status.name

                resolution = issue.fields.resolution
                if resolution is None:
                    result_data['解决结果'] = '未解决'
                else:
                    result_data['解决结果'] = resolution.name
                result_data['创建'] = issue.fields.created
                result_data['更新'] = issue.fields.updated

                component = issue.fields.components
                if len(component) > 0:
                    result_data['模块'] = component[0].name
                else:
                    result_data['模块'] = str(component)

                DI = issue.fields.customfield_12202
                if DI is None:
                    result_data['DI统计'] = DI
                else:
                    result_data['DI统计'] = DI.value

                tag = issue.fields.customfield_12202
                if tag is None:
                    result_data['标签'] = tag
                else:
                    result_data['标签'] = tag.value

                result_data['产品机型'] = issue.fields.customfield_10900
                result['data'].append(result_data)
                # 运营问题
                if result_data['模块'] == '应用商店':
                    store_data = {}
                    store_data['项目'] = projectId
                    store_data['问题类型'] = str(issue.fields.issuetype)
                    store_data['标识'] = issue.key
                    store_data['主题'] = issue.fields.summary
                    store_data['经办人'] = issue.fields.assignee.displayName
                    store_data['报告人'] = issue.fields.reporter.displayName
                    store_data['优先级'] = issue.fields.priority.name
                    store_data['状态'] = issue.fields.status.name

                    resolution = issue.fields.resolution
                    if resolution is None:
                        store_data['解决结果'] = '未解决'
                    else:
                        store_data['解决结果'] = resolution.name
                    store_data['创建'] = issue.fields.created
                    store_data['更新'] = issue.fields.updated

                    component = issue.fields.components
                    if len(component) > 0:
                        store_data['模块'] = component[0].name
                    else:
                        store_data['模块'] = str(component)

                    DI = issue.fields.customfield_12202
                    if DI is None:
                        store_data['DI统计'] = DI
                    else:
                        store_data['DI统计'] = DI.value

                    tag = issue.fields.customfield_12202
                    if tag is None:
                        store_data['标签'] = tag
                    else:
                        store_data['标签'] = tag.value

                    store_data['产品机型'] = issue.fields.customfield_10900
                    store_list.append(store_data)
                # UE问题
                if result_data['问题类型'] == 'UE Bug':
                    ue_data = {}
                    ue_data['项目'] = projectId
                    ue_data['问题类型'] = str(issue.fields.issuetype)
                    ue_data['标识'] = issue.key
                    ue_data['主题'] = issue.fields.summary
                    ue_data['经办人'] = issue.fields.assignee.displayName
                    ue_data['报告人'] = issue.fields.reporter.displayName
                    ue_data['优先级'] = issue.fields.priority.name
                    ue_data['状态'] = issue.fields.status.name

                    resolution = issue.fields.resolution
                    if resolution is None:
                        ue_data['解决结果'] = '未解决'
                    else:
                        ue_data['解决结果'] = resolution.name
                    ue_data['创建'] = issue.fields.created
                    ue_data['更新'] = issue.fields.updated

                    component = issue.fields.components
                    if len(component) > 0:
                        ue_data['模块'] = component[0].name
                    else:
                        ue_data['模块'] = str(component)

                    DI = issue.fields.customfield_12202
                    if DI is None:
                        ue_data['DI统计'] = DI
                    else:
                        ue_data['DI统计'] = DI.value

                    tag = issue.fields.customfield_12202
                    if tag is None:
                        ue_data['标签'] = tag
                    else:
                        ue_data['标签'] = tag.value

                    ue_data['产品机型'] = issue.fields.customfield_10900
                    ue_list.append(ue_data)

            download_id = time.strftime("%Y%m%d%H%M%S", time.localtime())
            save_excel(download_id, result['data'], store_list, ue_list)
            result['success'] = True
            result['message'] = 'test data'
            result['downloadId'] = download_id
            return HttpResponse(json.dumps(result))
        except JIRAError as e:
            print(e)
            result['message'] = 'JIRA登录出错'
            return HttpResponse(json.dumps(result))
    else:
        return render(request, 'jira/jira-search-data.html')


def save_excel(download_id, data, data_store, data_ue):
    data_list = []
    data_store_list = []
    data_ue_list = []
    title = {'项目': '项目', '经办人': '经办人', '标识': '标识', '状态': '状态', '模块': '模块', '创建': '创建', '问题类型': '问题类型', '优先级': '优先级',
         '解决结果': '解决结果', '报告人': '报告人', 'DI统计': 'DI统计', '标签': '标签', '主题': '主题', '产品机型': '产品机型;', '更新': '更新'}
    data_list.append(title)
    data_list.extend(data)
    # 运营问题
    data_store_list.append(title)
    data_store_list.extend(data_store)
    # UE问题
    data_ue_list.append(title)
    data_ue_list.extend(data_ue)

    # wb = Workbook()
    dest_filename = "/var/Data/JiraFile/" + download_id + u"软件成熟度评估报告.xlsx"
    # dest_filename = dest_filename.decode('utf-8')
    source_filename = settings.BASE_DIR + "/djangoWeb/static/JiraModel.xlsx"
    print(dest_filename)
    wb = load_workbook(source_filename)
    ws2 = wb["遗留问题列表"]
    ws3 = wb["运营问题"]
    ws4 = wb["UE问题"]

    # 遗留问题 Sheet
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)
    ws2['C1'].font = Font(bold=True)
    ws2['D1'].font = Font(bold=True)
    ws2['E1'].font = Font(bold=True)
    ws2['F1'].font = Font(bold=True)
    ws2['G1'].font = Font(bold=True)
    ws2['H1'].font = Font(bold=True)
    ws2['I1'].font = Font(bold=True)
    ws2['J1'].font = Font(bold=True)
    ws2['K1'].font = Font(bold=True)
    ws2['L1'].font = Font(bold=True)
    ws2['M1'].font = Font(bold=True)
    ws2['N1'].font = Font(bold=True)
    ws2['O1'].font = Font(bold=True)
    ws2.column_dimensions['D'].width = 25
    ws2.column_dimensions['J'].width = 15
    ws2.column_dimensions['K'].width = 15
    ws2.column_dimensions['O'].width = 15
    alg_lc = Alignment(horizontal='left', \
                       vertical='center', \
                       text_rotation=0, \
                       wrap_text=True, \
                       shrink_to_fit=True, \
                       indent=0)
    for row in range(1, len(data_list) + 1):
        for col in range(1, 16):
            ws2.cell(row=row, column=col).alignment = alg_lc
            if col == 1:
                ws2.cell(row=row, column=col, value=data_list[row - 1]['项目'])
            elif col == 2:
                ws2.cell(row=row, column=col, value=data_list[row - 1]['问题类型'])
            elif col == 3:
                ws2.cell(row=row, column=col, value=data_list[row - 1]['标识'])
            elif col == 4:
                ws2.cell(row=row, column=col, value=data_list[row - 1]['主题'])
            elif col == 5:
                ws2.cell(row=row, column=col, value=data_list[row - 1]['经办人'])
            elif col == 6:
                ws2.cell(row=row, column=col, value=data_list[row - 1]['报告人'])
            elif col == 7:
                ws2.cell(row=row, column=col, value=data_list[row - 1]['优先级'])
            elif col == 8:
                ws2.cell(row=row, column=col, value=data_list[row - 1]['状态'])
            elif col == 9:
                ws2.cell(row=row, column=col, value=data_list[row - 1]['解决结果'])
            elif col == 10:
                ws2.cell(row=row, column=col, value=data_list[row - 1]['创建'])
            elif col == 11:
                ws2.cell(row=row, column=col, value=data_list[row - 1]['更新'])
            elif col == 12:
                ws2.cell(row=row, column=col, value=data_list[row - 1]['模块'])
            elif col == 13:
                ws2.cell(row=row, column=col, value=data_list[row - 1]['DI统计'])
            elif col == 14:
                ws2.cell(row=row, column=col, value=data_list[row - 1]['标签'])
            else:
                ws2.cell(row=row, column=col, value=data_list[row - 1]['产品机型'])

    # 运营问题 Sheet
    ws3['A1'].font = Font(bold=True)
    ws3['B1'].font = Font(bold=True)
    ws3['C1'].font = Font(bold=True)
    ws3['D1'].font = Font(bold=True)
    ws3['E1'].font = Font(bold=True)
    ws3['F1'].font = Font(bold=True)
    ws3['G1'].font = Font(bold=True)
    ws3['H1'].font = Font(bold=True)
    ws3['I1'].font = Font(bold=True)
    ws3['J1'].font = Font(bold=True)
    ws3['K1'].font = Font(bold=True)
    ws3['L1'].font = Font(bold=True)
    ws3['M1'].font = Font(bold=True)
    ws3['N1'].font = Font(bold=True)
    ws3['O1'].font = Font(bold=True)
    ws3.column_dimensions['D'].width = 25
    ws3.column_dimensions['J'].width = 15
    ws3.column_dimensions['K'].width = 15
    ws3.column_dimensions['O'].width = 15
    alg_lc = Alignment(horizontal='left', \
                       vertical='center', \
                       text_rotation=0, \
                       wrap_text=True, \
                       shrink_to_fit=True, \
                       indent=0)
    for row in range(1, len(data_store_list) + 1):
        for col in range(1, 16):
            ws3.cell(row=row, column=col).alignment = alg_lc
            if col == 1:
                ws3.cell(row=row, column=col, value=data_store_list[row - 1]['项目'])
            elif col == 2:
                ws3.cell(row=row, column=col, value=data_store_list[row - 1]['问题类型'])
            elif col == 3:
                ws3.cell(row=row, column=col, value=data_store_list[row - 1]['标识'])
            elif col == 4:
                ws3.cell(row=row, column=col, value=data_store_list[row - 1]['主题'])
            elif col == 5:
                ws3.cell(row=row, column=col, value=data_store_list[row - 1]['经办人'])
            elif col == 6:
                ws3.cell(row=row, column=col, value=data_store_list[row - 1]['报告人'])
            elif col == 7:
                ws3.cell(row=row, column=col, value=data_store_list[row - 1]['优先级'])
            elif col == 8:
                ws3.cell(row=row, column=col, value=data_store_list[row - 1]['状态'])
            elif col == 9:
                ws3.cell(row=row, column=col, value=data_store_list[row - 1]['解决结果'])
            elif col == 10:
                ws3.cell(row=row, column=col, value=data_store_list[row - 1]['创建'])
            elif col == 11:
                ws3.cell(row=row, column=col, value=data_store_list[row - 1]['更新'])
            elif col == 12:
                ws3.cell(row=row, column=col, value=data_store_list[row - 1]['模块'])
            elif col == 13:
                ws3.cell(row=row, column=col, value=data_store_list[row - 1]['DI统计'])
            elif col == 14:
                ws3.cell(row=row, column=col, value=data_store_list[row - 1]['标签'])
            else:
                ws3.cell(row=row, column=col, value=data_store_list[row - 1]['产品机型'])
    # UE问题 Sheet
    ws4['A1'].font = Font(bold=True)
    ws4['B1'].font = Font(bold=True)
    ws4['C1'].font = Font(bold=True)
    ws4['D1'].font = Font(bold=True)
    ws4['E1'].font = Font(bold=True)
    ws4['F1'].font = Font(bold=True)
    ws4['G1'].font = Font(bold=True)
    ws4['H1'].font = Font(bold=True)
    ws4['I1'].font = Font(bold=True)
    ws4['J1'].font = Font(bold=True)
    ws4['K1'].font = Font(bold=True)
    ws4['L1'].font = Font(bold=True)
    ws4['M1'].font = Font(bold=True)
    ws4['N1'].font = Font(bold=True)
    ws4['O1'].font = Font(bold=True)
    ws4.column_dimensions['D'].width = 25
    ws4.column_dimensions['J'].width = 15
    ws4.column_dimensions['K'].width = 15
    ws4.column_dimensions['O'].width = 15
    alg_lc = Alignment(horizontal='left', \
                       vertical='center', \
                       text_rotation=0, \
                       wrap_text=True, \
                       shrink_to_fit=True, \
                       indent=0)
    for row in range(1, len(data_ue_list) + 1):
        for col in range(1, 16):
            ws4.cell(row=row, column=col).alignment = alg_lc
            if col == 1:
                ws4.cell(row=row, column=col, value=data_ue_list[row - 1]['项目'])
            elif col == 2:
                ws4.cell(row=row, column=col, value=data_ue_list[row - 1]['问题类型'])
            elif col == 3:
                ws4.cell(row=row, column=col, value=data_ue_list[row - 1]['标识'])
            elif col == 4:
                ws4.cell(row=row, column=col, value=data_ue_list[row - 1]['主题'])
            elif col == 5:
                ws4.cell(row=row, column=col, value=data_ue_list[row - 1]['经办人'])
            elif col == 6:
                ws4.cell(row=row, column=col, value=data_ue_list[row - 1]['报告人'])
            elif col == 7:
                ws4.cell(row=row, column=col, value=data_ue_list[row - 1]['优先级'])
            elif col == 8:
                ws4.cell(row=row, column=col, value=data_ue_list[row - 1]['状态'])
            elif col == 9:
                ws4.cell(row=row, column=col, value=data_ue_list[row - 1]['解决结果'])
            elif col == 10:
                ws4.cell(row=row, column=col, value=data_ue_list[row - 1]['创建'])
            elif col == 11:
                ws4.cell(row=row, column=col, value=data_ue_list[row - 1]['更新'])
            elif col == 12:
                ws4.cell(row=row, column=col, value=data_ue_list[row - 1]['模块'])
            elif col == 13:
                ws4.cell(row=row, column=col, value=data_ue_list[row - 1]['DI统计'])
            elif col == 14:
                ws4.cell(row=row, column=col, value=data_ue_list[row - 1]['标签'])
            else:
                ws4.cell(row=row, column=col, value=data_ue_list[row - 1]['产品机型'])

    ws2.auto_filter.ref = "A1:O1"  # 为表头增加过滤功能
    ws3.auto_filter.ref = "A1:O1"  # 为表头增加过滤功能
    ws4.auto_filter.ref = "A1:O1"  # 为表头增加过滤功能

    wb.save(filename=dest_filename)
    # response = HttpResponse(content_type='application/vnd.ms-excel;charset=utf-8;name="' + dest_filename + '"')
    # response['Content-Disposition'] = 'attachment; filename=ttttt'
    # # 保存返回
    # wb.save(response)
    # return response


def file_iterator(file, chunk_size=512):
    with open(file, 'rb') as f:
        while True:
            c = f.read(chunk_size)
            if c:
                yield c
            else:
                break


# 文件下载
def file_download(request):
    # do something...
    # file = "/root/JiraFile/www.xls"
    download_id = request.GET.get('downloadId', '')
    if download_id != "":
        file = "/var/Data/JiraFile/" + download_id + u"软件成熟度评估报告.xlsx"
        # file = unicode(file, "GB2312")
        file_name = file.split('/')[-1]
        print(file_name)
        response = StreamingHttpResponse(file_iterator(file))
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="{0}"'.format(urlquote(file_name))
        return response
    else:
        return HttpResponse("要下载的文件不存在！！！")



